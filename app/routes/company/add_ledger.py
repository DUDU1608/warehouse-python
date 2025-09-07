# app/routes/company/loan_ledger.py
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from sqlalchemy import asc

from app import db
from app.models import LoanLedger

bp = Blueprint("loan_ledger", __name__, url_prefix="/company/finance/loan-ledger")

# ---------- Helpers ----------
def _parse_date(s: str):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _compute_running_rows(rows):
    running = 0.0
    out = []
    for r in rows:
        running += (r.credit or 0.0) - (r.debit or 0.0)
        out.append({
            "id": r.id,
            "date": r.date,
            "purpose": r.purpose,
            "debit": float(r.debit or 0.0),
            "credit": float(r.credit or 0.0),
            "balance": float(running),
            "comments": r.comments or ""
        })
    return out

def _query_rows_with_filters():
    start = _parse_date(request.args.get("start", "") or request.form.get("start", "") or "")
    end   = _parse_date(request.args.get("end", "") or request.form.get("end", "") or "")

    q = LoanLedger.query
    if start:
        q = q.filter(LoanLedger.date >= start)
    if end:
        q = q.filter(LoanLedger.date <= end)
    rows = q.order_by(asc(LoanLedger.date), asc(LoanLedger.id)).all()
    return rows, start, end

# ---------- Pages ----------
@bp.route("/", methods=["GET", "POST"])
def index():
    # Add entry
    if request.method == "POST" and request.form.get("_action") == "create":
        date = _parse_date(request.form.get("date", "") or "")
        purpose = (request.form.get("purpose") or "").strip()
        debit = float(request.form.get("debit") or 0)
        credit = float(request.form.get("credit") or 0)
        comments = (request.form.get("comments") or "").strip()

        if not date:
            flash("Please provide a valid Date.", "danger")
            return redirect(url_for("loan_ledger.index"))
        if not purpose:
            flash("Purpose is required.", "danger")
            return redirect(url_for("loan_ledger.index"))
        if debit < 0 or credit < 0:
            flash("Debit/Credit cannot be negative.", "danger")
            return redirect(url_for("loan_ledger.index"))
        if debit == 0 and credit == 0:
            flash("Enter either Debit or Credit amount.", "warning")
            return redirect(url_for("loan_ledger.index"))

        entry = LoanLedger(date=date, purpose=purpose, debit=debit, credit=credit, comments=comments)
        db.session.add(entry)
        db.session.commit()
        flash("Loan ledger entry added.", "success")
        return redirect(url_for("loan_ledger.index"))

    # List with optional filters
    rows, start, end = _query_rows_with_filters()
    computed = _compute_running_rows(rows)
    totals = {
        "debit": sum(r["debit"] for r in computed),
        "credit": sum(r["credit"] for r in computed),
        "closing": computed[-1]["balance"] if computed else 0.0
    }
    return render_template("company/loan_ledger.html", rows=computed, totals=totals, start=start, end=end)

# ---------- Edit / Update ----------
@bp.route("/<int:row_id>/update", methods=["POST"])
def update_row(row_id: int):
    rec = LoanLedger.query.get_or_404(row_id)

    # Parse fields
    date = _parse_date(request.form.get("date", "") or "")
    purpose = (request.form.get("purpose") or "").strip()
    debit = float(request.form.get("debit") or 0)
    credit = float(request.form.get("credit") or 0)
    comments = (request.form.get("comments") or "").strip()

    # Basic validations
    if not date:
        flash("Invalid date.", "danger")
        return redirect(url_for("loan_ledger.index"))
    if not purpose:
        flash("Purpose is required.", "danger")
        return redirect(url_for("loan_ledger.index"))
    if debit < 0 or credit < 0:
        flash("Debit/Credit cannot be negative.", "danger")
        return redirect(url_for("loan_ledger.index"))
    if debit == 0 and credit == 0:
        flash("Enter either Debit or Credit amount.", "warning")
        return redirect(url_for("loan_ledger.index"))

    # Update
    rec.date = date
    rec.purpose = purpose
    rec.debit = debit
    rec.credit = credit
    rec.comments = comments
    db.session.commit()
    flash("Entry updated.", "success")
    return redirect(url_for("loan_ledger.index",
                            start=request.form.get("start") or "",
                            end=request.form.get("end") or ""))

# ---------- Delete ----------
@bp.route("/<int:row_id>/delete", methods=["POST"])
def delete_row(row_id: int):
    rec = LoanLedger.query.get_or_404(row_id)
    db.session.delete(rec)
    db.session.commit()
    flash("Entry deleted.", "success")
    return redirect(url_for("loan_ledger.index",
                            start=request.form.get("start") or "",
                            end=request.form.get("end") or ""))

# ---------- PDF ----------
@bp.route("/pdf", methods=["GET"])
def download_pdf():
    rows, start, end = _query_rows_with_filters()
    computed = _compute_running_rows(rows)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # --- Register Unicode font (₹ in headings) ---
    font_path = os.path.join(os.path.dirname(__file__),
                             "../../static/fonts/DejaVuSans.ttf")
    font_path = os.path.abspath(font_path)
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
        base_font = "DejaVuSans"
    else:
        base_font = "Helvetica"

    buffer = BytesIO()
    LEFT = RIGHT = 14 * mm
    TOP = BOTTOM = 16 * mm
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=LEFT, rightMargin=RIGHT,
        topMargin=TOP, bottomMargin=BOTTOM,
        title="Loan Ledger",
    )

    styles = getSampleStyleSheet()
    styles["Title"].fontName = base_font
    styles["Normal"].fontName = base_font

    body9 = ParagraphStyle(
        "body9",
        parent=styles["Normal"],
        fontName=base_font,
        fontSize=9,
        leading=11,
    )
    body9_bold = ParagraphStyle(
        "body9_bold",
        parent=body9,
        fontName=base_font,
        fontSize=9,
        leading=11,
    )

    story = []
    story.append(Paragraph("Loan Ledger (Company → Finance)", styles["Title"]))

    # Date range subtitle
    if start and end:
        subtitle_text = f"Period: {start.strftime('%d-%m-%Y')} to {end.strftime('%d-%m-%Y')}"
    elif start:
        subtitle_text = f"From: {start.strftime('%d-%m-%Y')}"
    elif end:
        subtitle_text = f"Upto: {end.strftime('%d-%m-%Y')}"
    else:
        subtitle_text = "All entries"
    story.append(Paragraph(subtitle_text, body9))
    story.append(Spacer(1, 6))

    # ---------- Column widths that FIT the page ----------
    page_width = A4[0]
    avail = page_width - (LEFT + RIGHT)  # printable width
    colw_mm = [20, 44, 30, 30, 28, 30]   # total = 182 mm
    colWidths = [w * mm for w in colw_mm]

    # ---------- Table data ----------
    data = [[
        Paragraph("Date", body9_bold),
        Paragraph("Purpose / Remarks", body9_bold),
        Paragraph("Debit (₹)", body9_bold),
        Paragraph("Credit (₹)", body9_bold),
        Paragraph("Balance (₹)", body9_bold),
        Paragraph("Comments", body9_bold),
    ]]

    for r in computed:
        data.append([
            Paragraph(r["date"].strftime("%d-%m-%Y"), body9),
            Paragraph(r["purpose"] or "", body9),
            Paragraph(f"{r['debit']:,.2f}", body9),
            Paragraph(f"{r['credit']:,.2f}", body9),
            Paragraph(f"{r['balance']:,.2f}", body9),
            Paragraph(r["comments"] or "", body9),
        ])

    # Totals row
    total_debit = sum(x["debit"] for x in computed)
    total_credit = sum(x["credit"] for x in computed)
    closing = computed[-1]["balance"] if computed else 0.0
    data.append([
        Paragraph("", body9),
        Paragraph("Totals", body9_bold),
        Paragraph(f"{total_debit:,.2f}", body9_bold),
        Paragraph(f"{total_credit:,.2f}", body9_bold),
        Paragraph(f"{closing:,.2f}", body9_bold),
        Paragraph("", body9),
    ])

    table = Table(data, colWidths=colWidths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, -1), base_font),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEADING", (0, 0), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("WORDWRAP", (1, 1), (1, -1), 1),
        ("WORDWRAP", (5, 1), (5, -1), 1),
        ("ALIGN", (2, 1), (4, -1), "RIGHT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F2F2F2")),
    ]))

    story.append(table)
    doc.build(story)

    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name="loan_ledger.pdf",
                     mimetype="application/pdf")

@bp.route("/excel", methods=["GET"])
def download_excel():
    # Use same filters
    rows, start, end = _query_rows_with_filters()
    computed = _compute_running_rows(rows)

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Ledger"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Loan Ledger (Company → Finance)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Subtitle (date range)
    ws.merge_cells("A2:F2")
    if start and end:
        sub = f"Period: {start.strftime('%d-%m-%Y')} to {end.strftime('%d-%m-%Y')}"
    elif start:
        sub = f"From: {start.strftime('%d-%m-%Y')}"
    elif end:
        sub = f"Upto: {end.strftime('%d-%m-%Y')}"
    else:
        sub = "All entries"
    ws["A2"] = sub
    ws["A2"].alignment = Alignment(horizontal="center")

    # Header row
    headers = ["Date", "Purpose / Remarks", "Debit (₹)", "Credit (₹)", "Balance (₹)", "Comments"]
    header_fill = PatternFill("solid", fgColor="EDEDED")
    header_font = Font(bold=True)
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Styles for numbers (Indian format if you like)
    # US/Intl style:
    money_fmt = "#,##0.00"
    # Indian lakh/crore style (works in Excel desktop):
    # money_fmt = '#,##,##0.00'

    money_style = NamedStyle(name="money_style")
    money_style.number_format = money_fmt
    money_style.alignment = Alignment(horizontal="right")

    # Data rows
    r0 = 4
    for i, r in enumerate(computed, start=0):
        row = r0 + i
        ws.cell(row=row, column=1, value=r["date"].strftime("%d-%m-%Y"))
        ws.cell(row=row, column=2, value=r["purpose"] or "")
        d = ws.cell(row=row, column=3, value=round(r["debit"], 2))
        c = ws.cell(row=row, column=4, value=round(r["credit"], 2))
        b = ws.cell(row=row, column=5, value=round(r["balance"], 2))
        ws.cell(row=row, column=6, value=r["comments"] or "")
        for cell in (d, c, b):
            cell.style = money_style

    # Totals row
    total_row = r0 + len(computed)
    ws.cell(row=total_row, column=2, value="Totals").font = Font(bold=True)
    td = ws.cell(row=total_row, column=3, value=round(sum(x["debit"] for x in computed), 2))
    tc = ws.cell(row=total_row, column=4, value=round(sum(x["credit"] for x in computed), 2))
    tb = ws.cell(row=total_row, column=5, value=round(computed[-1]["balance"], 2) if computed else 0.0)
    for cell in (td, tc, tb):
        cell.style = money_style
        cell.font = Font(bold=True)
    # Shade totals row
    for col in range(1, 7):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="F2F2F2")

    # Column widths (roughly matching PDF layout)
    widths = [12, 42, 16, 16, 18, 28]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Freeze panes (keep headers visible)
    ws.freeze_panes = "A4"

    # Output
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "loan_ledger.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

